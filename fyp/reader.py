from openpyxl import load_workbook
from threading import Thread

class Reader:
    def __init__(self):
        self.wb=load_workbook('static/record/template.xlsx')
        self.data = self.wb.get_sheet_by_name('data')



    def getNumber(self,type):
        number=-1
        if type=='Rent':
            number=0
        elif type == 'Sale':
            number=1
        elif type == 'Country Code':
            number=2
        elif type == 'Mobile Code':
            number=3
        elif type=='Numbering':
            number=4
        return number


    def getData(self,type):
        response=[]
        num=self.getNumber(type)
        if num>-1:
            counter=0
            for each_item in self.data.iter_rows():
                if each_item[num].value=='' or each_item[num].value=='null' or each_item[num].value==None:

                    break
                else:
                    if counter==0:
                        counter+=1
                    else:
                        response.append(each_item[num].value)

        return response


    def getPrice(self):
        response = []
        num ,counter= 0,0
        for each_item in self.data.iter_rows():
            if each_item[num].value == '' or each_item[num].value == 'null' or each_item[num].value == None:
                break
            else:
                if counter == 0:
                    counter += 1
                else:
                    response.append(each_item[num].value)
        return response

    def addComma(self,string):
        string=str(string)
        length = len(string) - 1
        resp, count = '', 0

        while (length > -1):
            if count == 3:
                resp += ','
                count = 0

            resp += string[length]
            count += 1
            length -= 1
        string = resp
        resp = ''
        length = len(string) - 1
        while (length > -1):
            resp += string[length]
            length -= 1
        return resp


    def execute_query(self,result_flag,cursor,query):
        flag,results=False,[]
        try:
            cursor.execute(query)
            flag=True
            if result_flag:
                results=cursor.fetchall()
        except:
            flag=False
        return flag,results

    def new_thread_for_contact(self,cron):
        th = Thread(cron.contactJob, ())
        th.deamon = True
        th.start()


    def new_thread_for_subscribe(self,cron, p_id):
        th = Thread(cron.subscriberJob, (p_id))
        th.deamon = True
        th.start()